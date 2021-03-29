from openpyxl import load_workbook
workbook = load_workbook(filename="PruebasMuestra.xlsx")
names=workbook.sheetnames
print(names)
print(workbook[names[0]])
sheet1=workbook[names[0]]
sheet2 = workbook[names[1]]
#print(sheet)
for row in sheet1.iter_rows(min_row=0,max_row=31870,min_col=0,max_col=4,values_only=True):
    print(row)




